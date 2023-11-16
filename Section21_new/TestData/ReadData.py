import openpyxl
# vk = openpyxl.load_workbook("D://TestData100.xlsx")
vk = openpyxl.load_workbook("C://coding/Learn_RF/Section21_new/TestData/TestData100.xlsx")


def fetch_number_of_rows(Sheetname):
    sh = vk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = vk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value


# print(fetch_number_of_rows("Taul1"))
# print(fetch_cell_data("Taul1", 1, 1))
# sh = vk["Taul1"]
# print(sh.max_row)
# cell = sh.cell(5,1)
# print(cell.value)

