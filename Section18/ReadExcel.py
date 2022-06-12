import openpyxl

# Load workbook

wk = openpyxl.load_workbook("c:\\temp\\testlog.xlsx")
# vk = openpyxl.load_workbook("testlog.xlsx")

print(wk.sheetnames)
print("Active sheet is: " + wk.active.title)

# Create an object on any sheet you want to work on

sh = wk['Sheet1']
print(sh.title)