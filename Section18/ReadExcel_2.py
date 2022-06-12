import openpyxl

# Load workbook

wk = openpyxl.load_workbook("c:\\temp\\testlog.xlsx")
sh = wk['Sheet1']

# print(sh['A3'].value)
# print(sh['B4'].value)

# c1 = sh.cell(3,1)
c1 = sh.cell(column=1, row=3)
print(c1.value)

print(c1.column)
print(c1.row)